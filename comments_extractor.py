import os
import pandas as pd
from openpyxl import load_workbook
import warnings

# Suppress data validation warnings
warnings.filterwarnings("ignore")

# Function to extract notes from a worksheet
def extract_notes(sheet, file_name, sheet_name):
    notes = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.comment:
                notes.append((file_name, sheet_name, cell.coordinate, cell.comment.text))
    return notes

# Directory containing Excel files
folder_path = r'C:\Users\user1\Desktop\excel files'

# List to store all notes
all_notes = []

# Loop through Excel files
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        
        # Read Excel file
        excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        
        # Extract notes from each sheet
        for sheet_name, sheet_data in excel_data.items():
            # Load worksheet using openpyxl for accessing comments
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            # Extract notes
            notes = extract_notes(ws, file_name, sheet_name)
            
            # Append notes to all_notes list
            all_notes.extend(notes)

# Create a DataFrame from the notes with filename first
notes_df = pd.DataFrame(all_notes, columns=['File Name', 'Sheet Name', 'Cell Coordinate', 'Note'])

# Specify the location to save the CSV file
csv_file_path = r'C:\Users\user1\Desktop\extracted_notes.csv'  # Change this path to your desired location

# Export DataFrame to CSV
notes_df.to_csv(csv_file_path, index=False)

print(f"Notes extracted and saved to '{csv_file_path}'")
